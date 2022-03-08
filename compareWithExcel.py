import openpyxl
import  tkinter
from tkinter import simpledialog
from tkinter import messagebox

def main():
    targetExcelFilePath = simpledialog.askstring("Input Box", "対象ファイルを指定して",)
    slicedTargetExcelFilePath = sliceStartAndEndCharacter(targetExcelFilePath)
    comparingExcelFilePath = simpledialog.askstring("Input Box", "比較するファイルを指定して",)
    slicedComparingExcelFilePath = sliceStartAndEndCharacter(comparingExcelFilePath)

    targetExcelBook  = openpyxl.load_workbook(slicedTargetExcelFilePath)
    targetExcelSheets = targetExcelBook.sheetnames
    comparingExcelBook = openpyxl.load_workbook(slicedComparingExcelFilePath)
    comparedExcelSheets = comparingExcelBook.sheetnames

    if targetExcelSheets == comparedExcelSheets:
        messagebox.showinfo('確認', '同じだから大丈夫')
    else:
        messagebox.showerror('エラー', 'シートが違うから確認して')

def sliceStartAndEndCharacter(targetString):
    slicedStartCharacterString = targetString[1:]
    slicedEndCharacterString = slicedStartCharacterString[:-1]
    return slicedEndCharacterString
